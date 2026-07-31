import openpyxl
import os
from datetime import datetime

def process_excel_file(file_path, output_file=None):
    processed_rows = 0
    
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        if output_file is None:
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_file = f"{base_name}_processed.txt"
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"# Обработано из: {os.path.basename(file_path)}\n")
            f.write(f"# Время обработки: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"# Всего строк: {sheet.max_row}\n")
            f.write("#" * 50 + "\n\n")
            
            for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                processed_cells = []
                for cell in row:
                    if cell is None:
                        processed_cells.append("''")
                    else:
                        processed_cells.append(f"'{str(cell).strip()}'")
                
                result = f"{','.join(processed_cells)},\n"
                f.write(result)
                processed_rows += 1
                
                # Прогресс каждые 100 строк
                if row_num % 100 == 0:
                    print(f"Обработано строк: {row_num}")
        
        return processed_rows, output_file
        
    except Exception as e:
        print(f"Ошибка: {e}")
        return 0, None

def main():
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    input_file = os.path.join(desktop_path, "data.xlsx")
    
    if not os.path.exists(input_file):
        print("Файл data.xlsx не найден на рабочем столе!")
        
        excel_files = [f for f in os.listdir(desktop_path) if f.endswith(('.xlsx', '.xls'))]
        if excel_files:
            print("Найдены другие Excel файлы:")
            for i, f in enumerate(excel_files, 1):
                print(f"{i}. {f}")
            
            try:
                choice = int(input("Выберите файл (номер): ")) - 1
                if 0 <= choice < len(excel_files):
                    input_file = os.path.join(desktop_path, excel_files[choice])
                else:
                    print("Неверный выбор!")
                    return
            except:
                print("Используется первый найденный файл")
                input_file = os.path.join(desktop_path, excel_files[0])
        else:
            print("Excel файлы не найдены на рабочем столе!")
            return
    
    output_file = input("Введите имя выходного файла (или Enter для автоматического): ").strip()
    if not output_file:
        output_file = None
    
    print(f"Обработка файла: {os.path.basename(input_file)}")
    
    start_time = datetime.now()
    processed_rows, output_path = process_excel_file(input_file, output_file)
    end_time = datetime.now()
    
    print("\n" + "="*50)
    print("РЕЗУЛЬТАТЫ ОБРАБОТКИ")
    print("="*50)
    print(f"✓ Обработано строк: {processed_rows}")
    print(f"✓ Входной файл: {os.path.basename(input_file)}")
    print(f"✓ Выходной файл: {output_path}")
    print(f"✓ Время обработки: {(end_time - start_time).total_seconds():.2f} сек")
    print(f"✓ Размер выходного файла: {os.path.getsize(output_path) if output_path and os.path.exists(output_path) else 0} байт")
    print("="*50)
    
    if processed_rows > 0:
        try:
            open_file = input("\nОткрыть полученный файл? (y/n): ").lower()
            if open_file == 'y':
                os.startfile(output_path)  # Для Windows
                print("Файл открыт!")
        except:
            print("Не удалось открыть файл автоматически")

if __name__ == "__main__":
    main()
    