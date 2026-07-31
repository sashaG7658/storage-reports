import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from config import CONFIG

def create_excel_report(products, ozon_accounts_items, wb_accounts_items):
    wb = Workbook()
    ws = wb.active
    ws.title = "Наличие товаров"
    
    header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    center_aligned = Alignment(horizontal='center')
    
    headers = ["Артикул"]
    
    for account in CONFIG['OZON_ACCOUNTS']:
        headers.append(f"Ozon: {account['name']}")
    
    for account in CONFIG['WILDBERRIES_ACCOUNTS']:
        headers.append(f"WB: {account['name']}")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_aligned
    
    for row_num, product in enumerate(products, 2):
        # Артикул
        ws.cell(row=row_num, column=1, value=product).border = border
        
        col_num = 2
        
        for account in CONFIG['OZON_ACCOUNTS']:
            account_name = account['name']
            status = ozon_accounts_items.get(account_name, {}).get(product, False)
            status_cell = ws.cell(row=row_num, column=col_num, value="+" if status else "-")
            status_cell.fill = PatternFill(start_color="C8E6C9" if status else "FFCDD2", fill_type="solid")
            status_cell.font = Font(bold=True)
            status_cell.alignment = center_aligned
            status_cell.border = border
            col_num += 1
        
        for account in CONFIG['WILDBERRIES_ACCOUNTS']:
            account_name = account['name']
            status = wb_accounts_items.get(account_name, {}).get(product, False)
            status_cell = ws.cell(row=row_num, column=col_num, value="+" if status else "-")
            status_cell.fill = PatternFill(start_color="C8E6C9" if status else "FFCDD2", fill_type="solid")
            status_cell.font = Font(bold=True)
            status_cell.alignment = center_aligned
            status_cell.border = border
            col_num += 1
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    try:
        desktop = os.path.join(os.path.expanduser("~"), "OneDrive", "Рабочий стол")
        if not os.path.exists(desktop):
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop, "marketplace_products_report_detailed.xlsx")
        wb.save(file_path)
    except Exception as e:
        file_path = "marketplace_products_report_detailed.xlsx"
        wb.save(file_path)
    
    return file_path